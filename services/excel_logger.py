# ============================================================================
# JIG ONE v1.2 - Excel Logger Service
# ============================================================================
# Monthly Excel files with daily subfolders
# Structure:
#   Offline_Logs/
#   ├── YYYY-MM/
#   │   ├── {PRODUCT}_{YYYY-MM}_monthly.xlsx
#   │   ├── YYYY-MM-DD/
#   │   │   └── {PRODUCT}_{YYYY-MM-DD}.xlsx
#   │   └── ...
#   ├── Pending_Upload/
#   │   ├── monthly/
#   │   └── daily/
#   └── S3_Uploaded/
#       ├── monthly/
#       └── daily/
# ============================================================================

import os
import shutil
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import OFFLINE_LOGS_DIR


class ExcelLogger:
    """
    Excel logging with monthly files and daily subfolders.
    
    Features:
    - One monthly file per month (rolling accumulation)
    - One daily file per day (same data)
    - Dynamic headers from product JSON configuration
    - Color-coded PASS/FAIL cells
    - Attempt tracking per row
    - Clock frequency Hz column
    - Automatic folder structure management
    - Pending upload handling on startup
    """
    
    # Cell colors
    PASS_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    
    # Column widths
    DEFAULT_WIDTH = 15
    NARROW_WIDTH = 12
    WIDE_WIDTH = 18
    
    def __init__(self, log_dir: str = None, product_name: str = "PRODUCT"):
        """
        Initialize Excel logger.
        
        Args:
            log_dir: Base directory for logs (default: Offline_Logs/)
            product_name: Product name for file naming (from JSON)
        """
        self.log_dir = log_dir or OFFLINE_LOGS_DIR
        self.product_name = self._sanitize_filename(product_name)
        
        self._headers: List[str] = []
        self._header_widths: Dict[str, int] = {}
        
        # Ensure base directories exist
        self._ensure_directories()
        
        # Handle pending uploads from previous day on startup
        self._handle_startup_uploads()
    
    def _sanitize_filename(self, name: str) -> str:
        """Sanitize product name for use in filenames"""
        # Replace spaces and special chars with underscores
        invalid_chars = '<>:"/\\|?*'
        result = name
        for char in invalid_chars:
            result = result.replace(char, '_')
        result = result.replace(' ', '_')
        return result
    
    def _ensure_directories(self):
        """Ensure all required directories exist"""
        dirs_to_create = [
            self.log_dir,
            os.path.join(self.log_dir, "Pending_Upload", "monthly"),
            os.path.join(self.log_dir, "Pending_Upload", "daily"),
            os.path.join(self.log_dir, "S3_Uploaded", "monthly"),
            os.path.join(self.log_dir, "S3_Uploaded", "daily"),
        ]
        
        for dir_path in dirs_to_create:
            os.makedirs(dir_path, exist_ok=True)
    
    def set_product_name(self, name: str):
        """Set product name (from JSON config)"""
        self.product_name = self._sanitize_filename(name)
        print(f"[EXCEL] Product name set: {self.product_name}")
    
    def set_headers(self, headers: List[str]):
        """
        Set column headers for the Excel file.
        Headers are dynamically built from product JSON.
        
        Args:
            headers: List of header names
        """
        # Fixed headers first
        self._headers = [
            "Date",
            "Time", 
            "Serial No.",
            "MAC ID",
            "Config State",
            "Attempt",
            "Clock Frequency Hz",  # Raw Hz value
        ]
        
        # Add dynamic headers from JSON
        self._headers.extend(headers)
        
        # Set column widths
        self._header_widths = {
            "Date": 15,
            "Time": 12,
            "Serial No.": 15,
            "MAC ID": 18,
            "Config State": 15,
            "Attempt": 10,
            "Clock Frequency Hz": 18,
        }
        
        # Default width for test result columns
        for h in headers:
            if h not in self._header_widths:
                # Narrower for zones, wider for named tests
                if h.startswith("Zone ") or h.startswith("Relay "):
                    self._header_widths[h] = self.NARROW_WIDTH
                else:
                    self._header_widths[h] = self.DEFAULT_WIDTH
        
        print(f"[EXCEL] Headers configured: {len(self._headers)} columns")
    
    # ========================================================================
    # PATH HELPERS
    # ========================================================================
    
    def _get_monthly_dir(self, date: datetime = None) -> str:
        """Get monthly directory path"""
        if date is None:
            date = datetime.now()
        month_str = date.strftime("%Y-%m")
        return os.path.join(self.log_dir, month_str)
    
    def _get_daily_dir(self, date: datetime = None) -> str:
        """Get daily directory path"""
        if date is None:
            date = datetime.now()
        month_dir = self._get_monthly_dir(date)
        day_str = date.strftime("%Y-%m-%d")
        return os.path.join(month_dir, day_str)
    
    def _get_monthly_filepath(self, date: datetime = None) -> str:
        """Get monthly Excel file path"""
        if date is None:
            date = datetime.now()
        month_dir = self._get_monthly_dir(date)
        month_str = date.strftime("%Y-%m")
        filename = f"{self.product_name}_{month_str}_monthly.xlsx"
        return os.path.join(month_dir, filename)
    
    def _get_daily_filepath(self, date: datetime = None) -> str:
        """Get daily Excel file path"""
        if date is None:
            date = datetime.now()
        daily_dir = self._get_daily_dir(date)
        day_str = date.strftime("%Y-%m-%d")
        filename = f"{self.product_name}_{day_str}.xlsx"
        return os.path.join(daily_dir, filename)
    
    # ========================================================================
    # WORKBOOK MANAGEMENT
    # ========================================================================
    
    def _get_or_create_workbook(self, filepath: str) -> openpyxl.Workbook:
        """Get existing workbook or create new one with headers"""
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        if os.path.isfile(filepath):
            try:
                return openpyxl.load_workbook(filepath)
            except Exception as e:
                print(f"[EXCEL] Error loading {filepath}: {e}, creating new")
        
        # Create new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Test Results"
        
        # Add headers
        if self._headers:
            sheet.append(self._headers)
            
            # Style header row
            for col in range(1, len(self._headers) + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.CENTER_ALIGN
                
                # Set column width
                header_name = self._headers[col - 1]
                width = self._header_widths.get(header_name, self.DEFAULT_WIDTH)
                col_letter = get_column_letter(col)
                sheet.column_dimensions[col_letter].width = width
            
            # Freeze header row
            sheet.freeze_panes = 'A2'
        
        return wb
    
    # ========================================================================
    # LOGGING
    # ========================================================================
    
    def log_result(self, result_dict: Dict[str, Any]) -> bool:
        """
        Log a test result to both daily and monthly Excel files.
        
        Args:
            result_dict: Dictionary with test results. Keys should match headers.
                         Special keys: '_id' (internal), 'clock_frequency_hz'
            
        Returns:
            True if logged successfully
        """
        try:
            now = datetime.now()
            
            # Ensure directories exist for current date
            os.makedirs(self._get_daily_dir(now), exist_ok=True)
            os.makedirs(self._get_monthly_dir(now), exist_ok=True)
            
            # Add date/time if not present
            if 'Date' not in result_dict:
                result_dict['Date'] = now.strftime('%d-%m-%Y')
            if 'Time' not in result_dict:
                result_dict['Time'] = now.strftime('%H:%M:%S')
            
            # Log to both files
            daily_success = self._log_to_file(self._get_daily_filepath(now), result_dict)
            monthly_success = self._log_to_file(self._get_monthly_filepath(now), result_dict)
            
            if daily_success and monthly_success:
                print(f"[EXCEL] Logged result to daily and monthly files")
                return True
            else:
                print(f"[EXCEL] Partial log failure: daily={daily_success}, monthly={monthly_success}")
                return False
            
        except Exception as e:
            print(f"[EXCEL] Error logging result: {e}")
            return False
    
    def _log_to_file(self, filepath: str, result_dict: Dict[str, Any]) -> bool:
        """
        Log result to a specific Excel file.
        
        Args:
            filepath: Path to Excel file
            result_dict: Dictionary with test results
            
        Returns:
            True if successful
        """
        try:
            wb = self._get_or_create_workbook(filepath)
            sheet = wb.active
            
            # Remove internal keys
            clean_dict = {k: v for k, v in result_dict.items() if not k.startswith('_')}
            
            # Build row values in header order
            row_values = []
            for header in self._headers:
                value = clean_dict.get(header, "")
                row_values.append(value)
            
            # Append row
            sheet.append(row_values)
            
            # Color code the new row
            last_row = sheet.max_row
            for col in range(1, len(self._headers) + 1):
                cell = sheet.cell(row=last_row, column=col)
                cell.alignment = self.CENTER_ALIGN
                
                if cell.value == "PASS":
                    cell.fill = self.PASS_FILL
                elif cell.value == "FAIL":
                    cell.fill = self.FAIL_FILL
            
            # Save
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"[EXCEL] Error writing to {filepath}: {e}")
            return False
    
    # ========================================================================
    # SEARCH & QUERY
    # ========================================================================
    
    def search_serial(self, serial_no: str, date: datetime = None) -> Optional[Dict[str, str]]:
        """
        Search for a serial number in today's daily log.
        
        Args:
            serial_no: Serial number to search for
            date: Date to search (default: today)
            
        Returns:
            Dictionary with result data, or None if not found
        """
        if date is None:
            date = datetime.now()
        
        filepath = self._get_daily_filepath(date)
        
        if not os.path.isfile(filepath):
            return None
        
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Find Serial No. column
            if "Serial No." not in headers:
                return None
            
            serial_col = headers.index("Serial No.")
            
            # Search for serial number (get last occurrence)
            last_match = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[serial_col]) == serial_no:
                    last_match = dict(zip(headers, row))
            
            return last_match
            
        except Exception as e:
            print(f"[EXCEL] Error searching: {e}")
            return None
    
    def get_serial_state(self, serial_no: str) -> str:
        """
        Get the Config State for a serial number from today's log.
        
        Args:
            serial_no: Serial number to check
            
        Returns:
            'SET', 'NOT_SET', or 'NOT_FOUND'
        """
        result = self.search_serial(serial_no)
        
        if result is None:
            return "NOT_FOUND"
        
        state = result.get("Config State", "")
        if state == "SET":
            return "SET"
        return "NOT_SET"
    
    def get_today_summary(self) -> Dict[str, int]:
        """
        Get summary of today's test results.
        
        Returns:
            Dictionary with counts of total, passed, failed tests
        """
        filepath = self._get_daily_filepath()
        
        summary = {"total": 0, "passed": 0, "failed": 0}
        
        if not os.path.isfile(filepath):
            return summary
        
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            
            # Find Config State column
            if "Config State" not in headers:
                return summary
            
            state_col = headers.index("Config State")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                summary["total"] += 1
                if row[state_col] == "SET":
                    summary["passed"] += 1
                else:
                    summary["failed"] += 1
            
            return summary
            
        except Exception as e:
            print(f"[EXCEL] Error getting summary: {e}")
            return summary
    
    def get_monthly_summary(self, date: datetime = None) -> Dict[str, int]:
        """
        Get summary of month's test results.
        
        Args:
            date: Any date in the target month (default: current month)
            
        Returns:
            Dictionary with counts
        """
        if date is None:
            date = datetime.now()
        
        filepath = self._get_monthly_filepath(date)
        
        summary = {"total": 0, "passed": 0, "failed": 0}
        
        if not os.path.isfile(filepath):
            return summary
        
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            
            if "Config State" not in headers:
                return summary
            
            state_col = headers.index("Config State")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                summary["total"] += 1
                if row[state_col] == "SET":
                    summary["passed"] += 1
                else:
                    summary["failed"] += 1
            
            return summary
            
        except Exception as e:
            print(f"[EXCEL] Error getting monthly summary: {e}")
            return summary
    
    # ========================================================================
    # UPLOAD HANDLING
    # ========================================================================
    
    def _handle_startup_uploads(self):
        """
        Handle pending uploads on application startup.
        
        Actions:
        1. Move previous day's daily file to Pending_Upload/daily/
        2. If first day of month, move previous month's file to Pending_Upload/monthly/
        """
        now = datetime.now()
        yesterday = now - timedelta(days=1)
        
        # 1. Check for yesterday's daily file
        self._move_daily_to_pending(yesterday)
        
        # 2. Check if it's first day of month
        if now.day == 1:
            # Move previous month's file to pending
            last_month = yesterday  # Yesterday is in previous month
            self._move_monthly_to_pending(last_month)
        
        print(f"[EXCEL] Startup upload check completed")
    
    def _move_daily_to_pending(self, date: datetime):
        """
        Move daily file(s) to Pending_Upload folder.
        
        Scans the daily directory for any .xlsx files instead of matching
        by exact product name. This ensures files are found regardless of
        the product_name set at init time vs. runtime.
        
        Args:
            date: Date of the daily file(s) to move
        """
        daily_dir = self._get_daily_dir(date)
        
        if not os.path.isdir(daily_dir):
            return
        
        # Scan for any .xlsx files in the daily directory
        xlsx_files = [f for f in os.listdir(daily_dir) if f.endswith('.xlsx')]
        
        if not xlsx_files:
            return
        
        # Create destination directory
        day_str = date.strftime("%Y-%m-%d")
        dest_dir = os.path.join(self.log_dir, "Pending_Upload", "daily", day_str)
        os.makedirs(dest_dir, exist_ok=True)
        
        for filename in xlsx_files:
            source = os.path.join(daily_dir, filename)
            dest = os.path.join(dest_dir, filename)
            
            # Skip if already copied
            if os.path.isfile(dest):
                print(f"[EXCEL] Daily file already in pending, skipping: {filename}")
                continue
            
            try:
                shutil.copy2(source, dest)
                print(f"[EXCEL] Copied daily file to pending: {dest}")
            except Exception as e:
                print(f"[EXCEL] Error copying daily to pending: {e}")
    
    def _move_monthly_to_pending(self, date: datetime):
        """
        Move monthly file(s) to Pending_Upload folder.
        
        Scans the monthly directory for any *_monthly.xlsx files instead of
        matching by exact product name. This ensures files are found regardless
        of the product_name set at init time vs. runtime.
        
        Args:
            date: Any date in the month to move
        """
        monthly_dir = self._get_monthly_dir(date)
        
        if not os.path.isdir(monthly_dir):
            return
        
        # Scan for any monthly .xlsx files in the monthly directory
        xlsx_files = [f for f in os.listdir(monthly_dir) 
                      if f.endswith('_monthly.xlsx')]
        
        if not xlsx_files:
            return
        
        dest_dir = os.path.join(self.log_dir, "Pending_Upload", "monthly")
        os.makedirs(dest_dir, exist_ok=True)
        
        for filename in xlsx_files:
            source = os.path.join(monthly_dir, filename)
            dest = os.path.join(dest_dir, filename)
            
            # Skip if already copied
            if os.path.isfile(dest):
                print(f"[EXCEL] Monthly file already in pending, skipping: {filename}")
                continue
            
            try:
                shutil.copy2(source, dest)
                print(f"[EXCEL] Copied monthly file to pending: {dest}")
            except Exception as e:
                print(f"[EXCEL] Error copying monthly to pending: {e}")
    
    def mark_daily_uploaded(self, date: datetime):
        """
        Mark a daily file as uploaded (move to S3_Uploaded).
        Called by external process after successful S3 upload.
        
        Args:
            date: Date of the uploaded file
        """
        day_str = date.strftime("%Y-%m-%d")
        source_dir = os.path.join(self.log_dir, "Pending_Upload", "daily", day_str)
        dest_dir = os.path.join(self.log_dir, "S3_Uploaded", "daily", day_str)
        
        if os.path.isdir(source_dir):
            os.makedirs(dest_dir, exist_ok=True)
            for filename in os.listdir(source_dir):
                src = os.path.join(source_dir, filename)
                dst = os.path.join(dest_dir, filename)
                shutil.move(src, dst)
            # Remove empty source directory
            try:
                os.rmdir(source_dir)
            except:
                pass
    
    def mark_monthly_uploaded(self, date: datetime):
        """
        Mark a monthly file as uploaded (move to S3_Uploaded).
        Called by external process after successful S3 upload.
        
        Args:
            date: Any date in the uploaded month
        """
        month_str = date.strftime("%Y-%m")
        filename = f"{self.product_name}_{month_str}_monthly.xlsx"
        
        source = os.path.join(self.log_dir, "Pending_Upload", "monthly", filename)
        dest_dir = os.path.join(self.log_dir, "S3_Uploaded", "monthly")
        
        if os.path.isfile(source):
            os.makedirs(dest_dir, exist_ok=True)
            dest = os.path.join(dest_dir, filename)
            shutil.move(source, dest)
    
    # ========================================================================
    # UTILITY METHODS
    # ========================================================================
    
    def get_pending_uploads(self) -> Dict[str, List[str]]:
        """
        Get list of files pending upload.
        
        Returns:
            Dictionary with 'daily' and 'monthly' file lists
        """
        result = {'daily': [], 'monthly': []}
        
        # Daily files
        daily_dir = os.path.join(self.log_dir, "Pending_Upload", "daily")
        if os.path.isdir(daily_dir):
            for day_folder in os.listdir(daily_dir):
                day_path = os.path.join(daily_dir, day_folder)
                if os.path.isdir(day_path):
                    for f in os.listdir(day_path):
                        if f.endswith('.xlsx'):
                            result['daily'].append(os.path.join(day_path, f))
        
        # Monthly files
        monthly_dir = os.path.join(self.log_dir, "Pending_Upload", "monthly")
        if os.path.isdir(monthly_dir):
            for f in os.listdir(monthly_dir):
                if f.endswith('.xlsx'):
                    result['monthly'].append(os.path.join(monthly_dir, f))
        
        return result
    
    def cleanup_old_uploaded(self, days_to_keep: int = 30):
        """
        Clean up old uploaded files.
        
        Args:
            days_to_keep: Number of days to keep uploaded files
        """
        cutoff = datetime.now() - timedelta(days=days_to_keep)
        
        uploaded_daily = os.path.join(self.log_dir, "S3_Uploaded", "daily")
        if os.path.isdir(uploaded_daily):
            for day_folder in os.listdir(uploaded_daily):
                try:
                    folder_date = datetime.strptime(day_folder, "%Y-%m-%d")
                    if folder_date < cutoff:
                        folder_path = os.path.join(uploaded_daily, day_folder)
                        shutil.rmtree(folder_path)
                        print(f"[EXCEL] Cleaned up old upload: {folder_path}")
                except ValueError:
                    pass  # Skip folders that don't match date format


# ============================================================================
# SINGLETON INSTANCE
# ============================================================================
excel_logger = ExcelLogger()
